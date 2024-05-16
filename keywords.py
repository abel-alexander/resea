import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import glob


def is_cell_green(cell):
    """Check if a cell is filled with a specific green color (FF00B050)."""
    fill = cell.fill
    print(f"Debug: cell fill: {fill}")
    if isinstance(fill, PatternFill):
        print("Debug: fill is a PatternFill")
        if fill.fgColor and fill.fgColor.rgb:
            color = fill.fgColor.rgb
            print(f"Debug: fgColor.rgb = {color}")
        elif fill.bgColor and fill.bgColor.rgb:
            color = fill.bgColor.rgb
            print(f"Debug: bgColor.rgb = {color}")
        else:
            print("Debug: No valid color found")
            return False

        # Check if the color matches the specific green color code
        return color == 'FF00B050'
    print("Debug: fill is not a PatternFill")
    return False


def compile_data(file_paths):
    """Compile data from multiple Excel files into two separate DataFrames based on 'Keep' column color."""
    df_green_list = []
    df_not_green_list = []

    for file_path in file_paths:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=False):  # Assuming the first row is header
            keyword = row[0].value
            frequency = row[1].value
            keep = row[2].value
            cell = row[2]

            row_data = {'Keyword': keyword, 'Frequency': frequency, 'Keep': keep}

            if is_cell_green(cell):
                df_green_list.append(row_data)
            else:
                df_not_green_list.append(row_data)

    df_green = pd.DataFrame(df_green_list)
    df_not_green = pd.DataFrame(df_not_green_list)

    return df_green, df_not_green


# Define the path to your Excel files (e.g., using glob to find all Excel files in a directory)
file_paths = glob.glob("path_to_your_excel_files/*.xlsx")

# Compile the data
df_green, df_not_green = compile_data(file_paths)

# Print the results or save them to a new Excel file
print("DataFrame with green 'Keep' cells:")
print(df_green)
print("\nDataFrame without green 'Keep' cells:")
print(df_not_green)

# Optionally, save to new Excel files
df_green.to_excel("green_keep.xlsx", index=False)
df_not_green.to_excel("not_green_keep.xlsx", index=False)
